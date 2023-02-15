import openpyxl


# Carga el archivo de trabajo
workbook = openpyxl.load_workbook("D:\ArgentinaPrograma\python\BARRANQUERAS A CONCEPCION.xlsx")

# Selecciona la hoja activa
hoja = workbook["Hoja1"]
#Crea un nuevo libro de Excel para guardar la fila
new_book=openpyxl.Workbook()
new_sheet=new_book.active
new_sheet.title="HOJA DE BARRANQUERAS"
# Itera sobre las filas en la hoja
for row in hoja.iter_rows(values_only=True):
    # Imprime los valores en cada celda de la fila
    if "BARRANQUERAS" in row:
        # Si se encuentra, agrega la fila al nuevo libro de Excel
        new_sheet.append(row)
# Guarda el nuevo libro de Excel
new_book.save("D:\ArgentinaPrograma\python\LIBRO DE BARRANQUERAS.xlsx")
# Cierra el libro original
workbook.close()
